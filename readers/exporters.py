import io
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from readers.loaders import sanitize_excel_value

def build_batch_excel(results: list, show_clean: bool = True) -> bytes:
    """
    Строит Excel с прогнозами для списка запчастей.
    """
    from forecasting.runner import MONTH_RU, build_result_summary, plot_forecast

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводная"
    ws_detail = wb.create_sheet("Детали")

    thin = Side(style="thin", color="D0D7DE")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_cell(ws, r, c, v, bold=False, bg=None, align="center", fmt=None, color="000000"):
        cell = ws.cell(row=r, column=c, value=sanitize_excel_value(v))
        cell.font = Font(name="Calibri", bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = brd
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if fmt:
            cell.number_format = fmt


    fc_months = results[0].fc_months
    month_labels = [f"{MONTH_RU[m]} {y}" for y, m in fc_months]

    # Сводная таблица
    summary_headers = [
    "№",
    "Артикул",
    "Номенклатура",
    "Список аналогов",
    "Конечный остаток",
    "Метод продаж",
    "Метод ремонта",
    ]

    for lbl in month_labels:
        summary_headers += [f"{lbl} Продажи", f"{lbl} Ремонт"]
    summary_headers += ["Итого продажи", "Итого ремонт", "Итого спрос", "Нужно заказать"]

    for ci, h in enumerate(summary_headers, 1):
        write_cell(ws_summary, 1, ci, h, bold=True, bg="1A2744", color="FFFFFF")

    for ri, result in enumerate(results, 2):
        bg = "FFFFFF" if ri % 2 == 0 else "F0F4FA"
        write_cell(ws_summary, ri, 1, ri - 1, bg=bg)
        write_cell(ws_summary, ri, 2, result.article, bg=bg)
        write_cell(ws_summary, ri, 3, result.nomenclature, bg=bg, align="left")
        write_cell(ws_summary, ri, 4, result.analogs, bg=bg, align="left")
        write_cell(ws_summary, ri, 5, round(float(result.ending_stock), 1), bg=bg, fmt="#,##0.0")
        write_cell(ws_summary, ri, 6, result.sale.method, bg=bg)
        write_cell(ws_summary, ri, 7, result.repair.method, bg=bg)

        base = 8

        for i in range(len(fc_months)):
            sv = round(float(result.sale.forecast.iloc[i]), 1)
            rv = round(float(result.repair.forecast.iloc[i]), 1)
            write_cell(ws_summary, ri, base, sv, bg="FFF3CD", color="B8520A", fmt="#,##0.0")
            write_cell(ws_summary, ri, base + 1, rv, bg="F4ECF7", color="5B2C6F", fmt="#,##0.0")
            base += 2
            
        summary = build_result_summary(result)
        write_cell(ws_summary, ri, base, summary["sale_total"], bg="FFF3CD", color="B8520A", bold=True, fmt="#,##0.0")
        write_cell(ws_summary, ri, base + 1, summary["repair_total"], bg="F4ECF7", color="5B2C6F", bold=True, fmt="#,##0.0")
        write_cell(ws_summary, ri, base + 2, summary["total_demand"], bold=True, fmt="#,##0.0")
        write_cell(ws_summary, ri, base + 3, summary["need_to_order"], bold=True, fmt="#,##0.0",
            bg="DCFCE7", color="166534")


    ws_summary.freeze_panes = "H2"

    # Детали
    detail_headers = ["Артикул", "Номенклатура", "Тип", "Метод", "Нули %"] + month_labels + ["Итого"]
    for ci, h in enumerate(detail_headers, 1):
        write_cell(ws_detail, 1, ci, h, bold=True, bg="1A2744", color="FFFFFF")

    current_row = 2

    for result in results:
        # Строка продаж
        zero_pct_sale = int(round(
            (result.sale.series_raw == 0).sum() / max(len(result.sale.series_raw), 1) * 100
        )) if result.sale.series_raw is not None else 0

        write_cell(ws_detail, current_row, 1, result.article)
        write_cell(ws_detail, current_row, 2, result.nomenclature, align="left")
        write_cell(ws_detail, current_row, 3, "Продажи", bg="FFF3CD", color="B8520A", bold=True)
        write_cell(ws_detail, current_row, 4, result.sale.method)
        write_cell(ws_detail, current_row, 5, f"{zero_pct_sale}%")

        sale_total = 0
        for i in range(len(fc_months)):
            val = round(float(result.sale.forecast.iloc[i]), 1)
            sale_total += val
            write_cell(ws_detail, current_row, 6 + i, val, bg="FFF3CD", color="B8520A", fmt="#,##0.0")
        write_cell(ws_detail, current_row, 6 + len(fc_months), round(sale_total, 1), bold=True, fmt="#,##0.0")
        current_row += 1

        # Строка ремонта
        zero_pct_repair = int(round(
            (result.repair.series_raw == 0).sum() / max(len(result.repair.series_raw), 1) * 100
        )) if result.repair.series_raw is not None else 0

        write_cell(ws_detail, current_row, 1, result.article)
        write_cell(ws_detail, current_row, 2, result.nomenclature, align="left")
        write_cell(ws_detail, current_row, 3, "Ремонт", bg="F4ECF7", color="5B2C6F", bold=True)
        write_cell(ws_detail, current_row, 4, result.repair.method)
        write_cell(ws_detail, current_row, 5, f"{zero_pct_repair}%")

        repair_total = 0
        for i in range(len(fc_months)):
            val = round(float(result.repair.forecast.iloc[i]), 1)
            repair_total += val
            write_cell(ws_detail, current_row, 6 + i, val, bg="F4ECF7", color="5B2C6F", fmt="#,##0.0")
        write_cell(ws_detail, current_row, 6 + len(fc_months), round(repair_total, 1), bold=True, fmt="#,##0.0")
        current_row += 1

        # График
        fig = plot_forecast(result, show_clean=show_clean, figsize=(14, 4))
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)

        img = XLImage(buf)
        img.width = 900
        img.height = 250
        ws_detail.row_dimensions[current_row].height = 190
        ws_detail.add_image(img, f"A{current_row}")
        current_row += 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()