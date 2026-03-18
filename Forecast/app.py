import io
import ast
import sys
import importlib
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ── Попытка импортировать пайплайн ───────────────────────────────────────────
try:
    import df_preprocessing as dp
    import initial_prep as ip
    from collections import defaultdict
    PIPELINE_AVAILABLE = True
except ImportError:
    PIPELINE_AVAILABLE = False

# ── Конфигурация страницы ────────────────────────────────────────────────────
st.set_page_config(
    page_title="Прогноз запчастей",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

.app-header {
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%);
    color: #e2e8f0; padding: 1.4rem 1.8rem; border-radius: 10px;
    margin-bottom: 1.2rem; border-left: 4px solid #3b82f6;
}
.app-header h1 { margin: 0; font-size: 1.5rem; font-weight: 600;
    font-family: 'IBM Plex Mono', monospace; letter-spacing: -0.5px; }
.app-header p { margin: 0.3rem 0 0; font-size: 0.82rem; color: #94a3b8; }

.metric-card { background: #f8fafc; border: 1px solid #e2e8f0;
    border-radius: 8px; padding: 0.9rem 1.1rem; text-align: center; }
.metric-card .val { font-size: 1.7rem; font-weight: 600;
    font-family: 'IBM Plex Mono', monospace; color: #1e40af; }
.metric-card .lbl { font-size: 0.75rem; color: #64748b; margin-top: 0.2rem; }

.badge { display: inline-block; padding: 2px 8px; border-radius: 4px;
    font-size: 0.72rem; font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
.badge-ets  { background: #dbeafe; color: #1d4ed8; }
.badge-cr   { background: #fce7f3; color: #9d174d; }
.badge-mean { background: #fef9c3; color: #854d0e; }

.fc-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
.fc-table th { background: #1e3a5f; color: #e2e8f0; padding: 0.5rem 0.7rem;
    text-align: center; font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; }
.fc-table td { padding: 0.45rem 0.7rem; border-bottom: 1px solid #f1f5f9; }
.fc-table tr:hover td { background: #f8fafc; }
.pos { color: #15803d; font-weight: 600; }
.neg { color: #dc2626; font-weight: 600; }

.sect-title { font-size: 0.78rem; font-weight: 600; letter-spacing: 1.5px;
    text-transform: uppercase; color: #94a3b8; border-bottom: 1px solid #e2e8f0;
    padding-bottom: 0.3rem; margin: 1rem 0 0.7rem; }

section[data-testid="stSidebar"] { background: #0f172a; }
section[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
</style>
""", unsafe_allow_html=True)


MONTH_RU = {1:'Янв',2:'Фев',3:'Мар',4:'Апр',5:'Май',6:'Июн',
            7:'Июл',8:'Авг',9:'Сен',10:'Окт',11:'Ноя',12:'Дек'}

def month_label(y, m):
    return f"{MONTH_RU[m]} {y}"

def next_months(year, month, n):
    result = []
    y, m = year, month
    for _ in range(n):
        m += 1
        if m > 12: m = 1; y += 1
        result.append((y, m))
    return result

def get_train_end(df):
    active = df[(df['Продажа'] > 0) | (df['Ремонт'] > 0)]
    if active.empty:
        return int(df['Год'].max()), int(df['Месяц'].max())
    last = active.sort_values(['Год','Месяц']).iloc[-1]
    return int(last['Год']), int(last['Месяц'])

@st.cache_data(show_spinner=False)
def load_final(file_bytes):
    return pd.read_excel(io.BytesIO(file_bytes))

def make_monthly(df):
    df = df.copy()
    df['date'] = pd.to_datetime(
        df['Год'].astype(str) + '-' + df['Месяц'].astype(str).str.zfill(2) + '-01')
    return df.groupby(['date','Номер группы'])[['Продажа','Ремонт']].sum().reset_index()

def get_series_dyn(monthly, grp, col, train_end_year, train_end_month):
    end = pd.Timestamp(year=train_end_year, month=train_end_month, day=1)
    full_dates = pd.date_range('2023-01-01', end, freq='MS')
    sub = monthly[monthly['Номер группы'] == grp][['date', col]].set_index('date')
    return sub[col].reindex(full_dates, fill_value=0).astype(float)

def remove_outliers_local(series, iqr_factor=1.5):
    s = series.copy()
    q1, q3 = s.quantile(0.25), s.quantile(0.75)
    iqr = q3 - q1
    upper = q3 + iqr_factor * iqr
    mask = s > upper
    log = []
    for idx in s[mask].index:
        orig = s[idx]; pos = s.index.get_loc(idx)
        nbrs = [s.iloc[pos+off] for off in [-3,-2,-1,1,2,3]
                if 0 <= pos+off < len(s) and not mask.iloc[pos+off]]
        repl = float(np.median(nbrs)) if nbrs else float(s.median())
        s.iloc[pos] = repl
        log.append({'date': idx.strftime('%Y-%m'), 'original': round(orig,1),
                    'replacement': round(repl,1), 'threshold': round(upper,1)})
    return s, log

def forecast_series(series, steps, fc_start_date):
    from statsmodels.tsa.holtwinters import ExponentialSmoothing as ETS
    s = series.clip(lower=0)
    zero_ratio = (s == 0).sum() / max(len(s), 1)
    idx = pd.date_range(fc_start_date, periods=steps, freq='MS')

    if zero_ratio > 0.40:
        vals = s.values; alpha = 0.1
        nz = [(i,v) for i,v in enumerate(vals) if v > 0]
        if not nz:
            return pd.Series([0.0]*steps, index=idx), 'Croston'
        z = nz[0][1]; p = float(nz[0][0]+1); prev = nz[0][0]
        for i2, v2 in nz[1:]:
            z = alpha*v2 + (1-alpha)*z
            p = alpha*(i2-prev) + (1-alpha)*p; prev = i2
        fc_val = max(0.0, round(z/p, 1)) if p > 0 else 0.0
        return pd.Series([fc_val]*steps, index=idx), 'Croston'

    has_zeros = (s == 0).any()
    candidates = [
        ('add','add',12,'HW-add'),('add','mul',12,'HW-mul'),
        ('add',None,None,'Holt'),(None,'add',12,'ETS-сез'),(None,None,None,'ETS'),
    ]
    best_fc, best_aic, best_name = None, np.inf, None
    for trend, seasonal, sp, name in candidates:
        if seasonal == 'mul' and has_zeros: continue
        if sp and len(s) < 2*sp: continue
        try:
            m = ETS(s, trend=trend, seasonal=seasonal, seasonal_periods=sp)
            f = m.fit(optimized=True)
            if f.aic < best_aic:
                best_aic = f.aic
                best_fc = f.forecast(steps).clip(lower=0).round(1)
                best_name = name
        except Exception:
            pass

    if best_fc is None:
        v = s.tail(6).mean()
        best_fc = pd.Series([round(v,1)]*steps, index=idx)
        best_name = 'Mean-6m'
    else:
        best_fc.index = idx
    return best_fc, best_name

def run_forecast(df, group_ids, steps, iqr_factor, col='Продажа'):
    train_end_year, train_end_month = get_train_end(df)
    monthly = make_monthly(df)
    fc_months = next_months(train_end_year, train_end_month, steps)
    fc_start  = pd.Timestamp(year=fc_months[0][0], month=fc_months[0][1], day=1)

    results = []
    for grp in group_ids:
        rows = df[df['Номер группы'] == grp]
        meta = rows.iloc[0] if len(rows) else None
        raw = get_series_dyn(monthly, grp, col, train_end_year, train_end_month)
        cleaned, outliers = remove_outliers_local(raw, iqr_factor)
        fc_raw,   method_raw   = forecast_series(raw,     steps, fc_start)
        fc_clean, method_clean = forecast_series(cleaned, steps, fc_start)
        fact = [round(float(df[(df['Номер группы']==grp)&(df['Год']==y)&(df['Месяц']==m)][col].sum()),1)
                for y, m in fc_months]
        results.append({
            'grp':          grp,
            'name':         str(meta['Номенклатура'])[:70] if meta is not None else '—',
            'article':      str(meta['Артикул']) if meta is not None else '—',
            'fc_months':    fc_months,
            'raw':          raw,
            'cleaned':      cleaned,
            'outliers':     outliers,
            'fc_raw':       fc_raw,
            'fc_clean':     fc_clean,
            'method_raw':   method_raw,
            'method_clean': method_clean,
            'fact':         fact,
            'zero_pct':     int(round((raw==0).sum()/max(len(raw),1)*100)),
        })
    return results, fc_months

def badge_html(method):
    cls = 'badge-cr' if 'Croston' in method else ('badge-mean' if 'Mean' in method else 'badge-ets')
    return f'<span class="badge {cls}">{method}</span>'

def delta_html(d):
    if d is None: return '—'
    cls = 'pos' if d >= 0 else 'neg'
    return f'<span class="{cls}">{("+" if d > 0 else "")}{d:.1f}</span>'

def make_chart(result, show_clean=True):
    r = result
    raw = r['raw']
    fc  = r['fc_clean'] if show_clean else r['fc_raw']
    fig = go.Figure()

    fig.add_trace(go.Scatter(x=raw.index, y=raw.values, mode='lines+markers',
        name='История', line=dict(color='#3b82f6', width=2), marker=dict(size=4)))

    if show_clean and r['outliers']:
        cl = r['cleaned']
        fig.add_trace(go.Scatter(x=cl.index, y=cl.values, mode='lines',
            name='После очистки', line=dict(color='#8b5cf6', width=1.5, dash='dot')))

    last_date = raw.index[-1]
    fc_x = [last_date] + list(fc.index)
    fc_y = [raw.values[-1]] + list(fc.values)
    method = r['method_clean'] if show_clean else r['method_raw']
    fig.add_trace(go.Scatter(x=fc_x, y=fc_y, mode='lines+markers',
        name=f'Прогноз ({method})',
        line=dict(color='#f97316', width=2.5, dash='dash'),
        marker=dict(size=6, symbol='diamond')))

    if any(v > 0 for v in r['fact']):
        fact_x = [pd.Timestamp(y, m, 1) for y, m in r['fc_months']]
        fig.add_trace(go.Scatter(x=fact_x, y=r['fact'], mode='lines+markers',
            name='Факт', line=dict(color='#10b981', width=2),
            marker=dict(size=7, symbol='circle-open')))

    for o in r['outliers']:
        fig.add_trace(go.Scatter(x=[pd.Timestamp(o['date']+'-01')], y=[o['original']],
            mode='markers', showlegend=False,
            marker=dict(size=10, color='#ef4444', symbol='x')))

    fig.update_layout(
        height=300, margin=dict(l=10,r=10,t=25,b=10),
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(248,250,252,1)',
        font=dict(family='IBM Plex Sans', size=11),
        legend=dict(orientation='h', y=-0.2, font=dict(size=10)),
        xaxis=dict(gridcolor='#e2e8f0'), yaxis=dict(gridcolor='#e2e8f0'),
        hovermode='x unified',
    )
    return fig

def build_excel_output(results_sale, results_repair, fc_months):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Продажи"
    ws2 = wb.create_sheet("Ремонт")

    def write_sheet(ws_out, results):
        thin = Side(style='thin', color='D0D7DE')
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def w(r, c, v, bold=False, bg=None, align='center', fmt=None, color='000000'):
            cell = ws_out.cell(row=r, column=c, value=v)
            cell.font = Font(name='Calibri', bold=bold, color=color, size=10)
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            cell.border = brd
            if bg: cell.fill = PatternFill('solid', fgColor=bg)
            if fmt: cell.number_format = fmt

        labels = [month_label(y,m) for y,m in fc_months]
        has_fact = any(any(v>0 for v in r['fact']) for r in results)

        headers = ['№','Группа','Артикул','Номенклатура','Метод','Нули %']
        for lbl in labels:
            headers += [f'{lbl} Прогноз', f'{lbl} Без выбр.']
            if has_fact: headers += [f'{lbl} Факт', f'{lbl} Δ']
        headers += ['Итого Прогноз', 'Итого Без выбр.']
        if has_fact: headers += ['Итого Факт', 'Итого Δ']

        for ci, h in enumerate(headers, 1):
            w(1, ci, h, bold=True, bg='1A2744', color='FFFFFF')
        ws_out.row_dimensions[1].height = 32

        col_w = [4,8,14,40,14,8] + ([12,14,12,10] if has_fact else [12,14]) * len(labels)
        col_w += [14,16] + ([14,10] if has_fact else [])
        for ci, cw in enumerate(col_w[:len(headers)], 1):
            ws_out.column_dimensions[get_column_letter(ci)].width = cw

        steps = len(fc_months)
        for ri, res in enumerate(results, 2):
            bg = 'FFFFFF' if ri % 2 == 0 else 'F0F4FA'
            ws_out.row_dimensions[ri].height = 20
            w(ri,1,ri-1,bg=bg); w(ri,2,res['grp'],bg=bg)
            w(ri,3,res['article'],bg=bg)
            w(ri,4,res['name'],bg=bg,align='left')
            w(ri,5,res['method_clean'],bg=bg)
            w(ri,6,f"{res['zero_pct']}%",bg=bg)

            sum_raw=sum_clean=sum_fact=0
            base = 7
            for i in range(steps):
                fc_r = round(res['fc_raw'].iloc[i],1)
                fc_c = round(res['fc_clean'].iloc[i],1)
                fact = res['fact'][i]
                sum_raw+=fc_r; sum_clean+=fc_c; sum_fact+=fact
                w(ri,base,   fc_r, bg='FFF3CD',color='B8520A',bold=True,fmt='#,##0.0')
                w(ri,base+1, fc_c, bg='F4ECF7',color='5B2C6F',bold=True,fmt='#,##0.0')
                if has_fact:
                    w(ri,base+2,fact,bg='D5F5E3',color='117A65',bold=True,fmt='#,##0.0')
                    d = round(fact-fc_c,1)
                    dv = f"{'+' if d>0 else ''}{d:.1f}"
                    w(ri,base+3,dv,bg=bg,color=('15803D' if d>=0 else 'DC2626'),bold=True)
                    base += 4
                else:
                    base += 2

            w(ri,base,  round(sum_raw,1),  bg='FFF3CD',color='B8520A',bold=True,fmt='#,##0.0')
            w(ri,base+1,round(sum_clean,1),bg='F4ECF7',color='5B2C6F',bold=True,fmt='#,##0.0')
            if has_fact:
                w(ri,base+2,round(sum_fact,1),bg='D5F5E3',color='117A65',bold=True,fmt='#,##0.0')
                qd = round(sum_fact-sum_clean,1)
                w(ri,base+3,f"{'+' if qd>0 else ''}{qd:.1f}",bg=bg,
                  color=('15803D' if qd>=0 else 'DC2626'),bold=True)

        ws_out.freeze_panes = 'G2'

    write_sheet(ws1, results_sale)
    write_sheet(ws2, results_repair)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()



st.markdown("""
<div class="app-header">
  <h1> Прогноз запасных частей</h1>
  <p>Загрузите данные → найдите запчасти → получите прогноз и Excel-отчёт</p>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### Источник данных")
    data_mode = st.radio("Режим", ["Готовый final.xlsx", "Два исходных отчёта"],
                         label_visibility="collapsed")
    df_main = None

    if data_mode == "Готовый final.xlsx":
        uf = st.file_uploader("Загрузите final.xlsx", type=['xlsx'])
        if uf:
            with st.spinner("Загружаем..."):
                df_main = load_final(uf.read())
            st.success(f"{len(df_main):,} строк")

    else:
        st.markdown("**Отчёт 1 — Списанные в ремонт**")
        f1 = st.file_uploader("repair.xlsx", type=['xlsx'], key='f1',
                               label_visibility="collapsed")
        st.markdown("**Отчёт 2 — Остатки и обороты**")
        f2 = st.file_uploader("stock.xlsx", type=['xlsx'], key='f2',
                               label_visibility="collapsed")

        if f1 and f2:
            if not PIPELINE_AVAILABLE:
                st.error("df_preprocessing.py и initial_prep.py не найдены рядом с app.py")
            elif st.button(" Обработать", type="primary"):
                with st.spinner("Обрабатываем (~1-2 мин)..."):
                    import tempfile, os
                    with tempfile.TemporaryDirectory() as tmp:
                        p1 = os.path.join(tmp,'repair.xlsx')
                        p2 = os.path.join(tmp,'stock.xlsx')
                        open(p1,'wb').write(f1.read())
                        open(p2,'wb').write(f2.read())
                        try:
                            # Повторяем пайплайн из ноутбука
                            data = dp.load_repair_parts(None, p1, 8)
                            df1  = ip.data_prep_1(data)
                            df1  = ip.del_wrong_numbers_df1(df1)
                            for c in ['Номенклатура.Артикул','Номенклатура.Оригинальный номер',
                                      'Номенклатура.Оригинальный номер расширенный']:
                                df1[c] = df1[c].apply(dp.normalize)
                            df1["Номенклатура.Оригинальный номер расширенный"] = df1.apply(
                                dp.add_original_to_extended, axis=1)
                            df1['Аналоги'] = (df1['Номенклатура.Оригинальный номер расширенный']
                                              .fillna('').str.upper().str.split())
                            graph = defaultdict(set)
                            for _, row in df1.iterrows():
                                pf = dp.article_forms(row['Номенклатура.Артикул'])
                                for i in range(len(pf)):
                                    for j in range(i+1,len(pf)):
                                        graph[pf[i]].add(pf[j]); graph[pf[j]].add(pf[i])
                                for ar in row['Аналоги']:
                                    for af in dp.article_forms(ar):
                                        for p in pf:
                                            if p!=af: graph[p].add(af); graph[af].add(p)
                            df1['all_analogs'] = df1['Номенклатура.Артикул'].apply(
                                lambda x: dp.find_all_analogs(x, graph))
                            df1 = df1.drop(columns='Аналоги')
                            gm={}; gc=1
                            for gt in df1['all_analogs'].unique(): gm[gt]=gc; gc+=1
                            df1['Номер группы'] = df1['all_analogs'].map(gm)
                            df_quarter = dp.merge_parts_df1(df1)

                            data2 = dp.load_repair_parts(None, p2, skiprows=8)
                            df2   = ip.data_prep_2(data2)
                            df2   = ip.del_wrong_numbers_df2(df2)
                            df2["Артикул"] = df2["Артикул"].apply(dp.normalize)
                            df2["Оригинальный номер"] = df2["Оригинальный номер"].apply(dp.normalize)
                            a2g={}; a2a={}
                            for _, row in df1.drop_duplicates('Номенклатура.Артикул').iterrows():
                                grp=row['Номер группы']; raw=row['all_analogs']
                                try: analogs=raw if isinstance(raw,tuple) else ast.literal_eval(raw)
                                except: analogs=()
                                for art in analogs: a2g[art]=grp; a2a[art]=analogs
                                main=row['Номенклатура.Артикул']
                                if pd.notna(main) and main: a2g[main]=grp; a2a[main]=analogs
                            def lookup(an, on):
                                for val in [an, on]:
                                    if not val: continue
                                    for form in dp.article_forms(val):
                                        if form in a2g: return a2g[form], a2a[form]
                                return None, None
                            groups, analogs_list = zip(*df2.apply(
                                lambda r: lookup(r["Артикул"],r["Оригинальный номер"]), axis=1))
                            df2["Номер группы"]=groups; df2["Список аналогов"]=analogs_list
                            def enrich(row):
                                art=row["Артикул"]; analogs=row["Список аналогов"]
                                if pd.isnull(art) or not isinstance(analogs,tuple): return analogs
                                if art not in analogs and art not in a2g:
                                    return tuple(sorted(analogs+(art,)))
                                return analogs
                            df2["Список аналогов"]=df2.apply(enrich,axis=1)
                            df2=dp.normalize_analog_lists(df2)
                            agg=dp.merge_parts_df2(df2)
                            agg=dp.prepare_sales(agg,df_quarter)
                            final=pd.merge(df_quarter,agg,on=['Год','Месяц','Номер группы'],
                                           how='outer',suffixes=('_df','_agg'))
                            cc=[c for c in df_quarter.columns if c in agg.columns
                                and c not in ['Год','Месяц','Номер группы']]
                            for col in cc:
                                final[col]=final[f'{col}_df'].combine_first(final[f'{col}_agg'])
                            final.drop(columns=[c for c in final.columns
                                                if c.endswith('_df') or c.endswith('_agg')],inplace=True)
                            final['Продажа']=final['Продажа'].fillna(0)
                            final['Ремонт']=final['Ремонт'].fillna(0)
                            final=dp.normalize_analog_lists(final,col_group="Номер группы",
                                                            col_analogs="Список аналогов")
                            nm=(final.groupby('Номер группы')['Номенклатура']
                                .apply(lambda s: min(s.dropna().astype(str),key=len,default=None)))
                            final['Номенклатура']=final['Номер группы'].map(nm)
                            final=dp.fill_missing_months(final)
                            df_main=final
                            st.session_state['df_main']=df_main
                            st.success("Готово!")
                        except Exception as e:
                            st.error(f"Ошибка: {e}")

    if 'df_main' in st.session_state and df_main is None:
        df_main = st.session_state['df_main']

    st.divider()
    st.markdown("### Параметры прогноза")
    steps      = st.slider("Горизонт (месяцев)", 1, 12, 3)
    show_clean = st.toggle("Прогноз без выбросов", value=True)
    iqr_factor = st.slider("Чувствительность к выбросам (IQR ×)", 1.0, 3.0, 1.5, 0.1)

# ── ОСНОВНАЯ ОБЛАСТЬ ─────────────────────────────────────────────────────────
if df_main is None:
    st.info("Загрузите данные в боковой панели")
    st.stop()

# Индекс для поиска
meta_df = (df_main.groupby('Номер группы')
           .agg(Номенклатура=('Номенклатура','first'),
                Артикул=('Артикул','first'),
                Продажа_сумм=('Продажа','sum'),
                Ремонт_сумм=('Ремонт','sum'))
           .reset_index().dropna(subset=['Номенклатура']))
meta_df['Номер группы'] = meta_df['Номер группы'].astype(int)
meta_df = meta_df.sort_values('Продажа_сумм', ascending=False)

# ── ПОИСК ────────────────────────────────────────────────────────────────────
st.markdown('<div class="sect-title">Поиск запчастей</div>', unsafe_allow_html=True)

col_s, col_t = st.columns([3, 1])
with col_s:
    query = st.text_input("🔍", placeholder="название или артикул...",
                          label_visibility="collapsed")
with col_t:
    top_n = st.selectbox("Топ", ["—","Топ 10","Топ 20","Топ 50"],
                         label_visibility="collapsed")

if query.strip():
    q = query.strip().lower()
    mask = (meta_df['Номенклатура'].str.lower().str.contains(q, na=False) |
            meta_df['Артикул'].str.lower().str.contains(q, na=False))
    filtered = meta_df[mask]
elif top_n != "—":
    filtered = meta_df.head(int(top_n.split()[1]))
else:
    filtered = meta_df.head(20)

if filtered.empty:
    st.warning("Ничего не найдено")
    st.stop()

st.caption(f"Найдено {len(filtered)} групп. Выберите для прогноза (макс. 30):")

options = {
    f"Гр.{int(r['Номер группы']):4d}  {str(r['Номенклатура'])[:55]}  [{str(r['Артикул'])[:15]}]": int(r['Номер группы'])
    for _, r in filtered.iterrows()
}

selected_labels = st.multiselect(
    "Группы", list(options.keys()),
    default=list(options.keys())[:min(5, len(options))],
    label_visibility="collapsed", max_selections=30,
)
selected_groups = [options[l] for l in selected_labels]

if not selected_groups:
    st.info("Выберите хотя бы одну группу")
    st.stop()

# ── КНОПКА ПРОГНОЗА ──────────────────────────────────────────────────────────
st.divider()
if st.button("▶ Построить прогноз", type="primary", use_container_width=True):
    with st.spinner("Считаем..."):
        r_sale,   fc_months = run_forecast(df_main, selected_groups, steps, iqr_factor, 'Продажа')
        r_repair, _         = run_forecast(df_main, selected_groups, steps, iqr_factor, 'Ремонт')
    st.session_state.update({'r_sale':r_sale,'r_repair':r_repair,'fc_months':fc_months})

if 'r_sale' not in st.session_state:
    st.stop()

r_sale   = st.session_state['r_sale']
r_repair = st.session_state['r_repair']
fc_months = st.session_state['fc_months']
fc_labels = [month_label(y,m) for y,m in fc_months]

# ── МЕТРИКИ ──────────────────────────────────────────────────────────────────
st.markdown('<div class="sect-title">Сводка</div>', unsafe_allow_html=True)
total_fc   = sum(sum(r['fc_clean'].values) for r in r_sale)
total_fact = sum(sum(r['fact']) for r in r_sale)
total_out  = sum(len(r['outliers']) for r in r_sale)

m1,m2,m3,m4 = st.columns(4)
for col_m, val, lbl in [
    (m1, len(selected_groups),         "Групп"),
    (m2, f"{total_fc:,.0f}",           f"Прогноз продаж ({len(fc_labels)} мес.)"),
    (m3, f"{total_fact:,.0f}" if total_fact>0 else "—", "Факт (если есть)"),
    (m4, total_out,                    "Выбросов"),
]:
    col_m.markdown(f'<div class="metric-card"><div class="val">{val}</div>'
                   f'<div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)

st.markdown("")

# ── ТАБЛИЦЫ + ГРАФИКИ ────────────────────────────────────────────────────────
tab_sale, tab_repair = st.tabs(["Продажи", "Ремонт"])

def render_tab(results, tab):
    with tab:
        has_fact = any(any(v>0 for v in r['fact']) for r in results)

        # Таблица
        hdr = "<th>Гр.</th><th>Артикул</th><th>Номенклатура</th><th>Метод</th><th>Нули</th>"
        for lbl in fc_labels:
            hdr += f"<th>{lbl}<br><small>Прогноз</small></th>"
            if has_fact:
                hdr += f"<th>{lbl}<br><small>Факт</small></th>"
                hdr += f"<th>{lbl}<br><small>Δ</small></th>"
        hdr += "<th>Итого<br><small>Прогноз</small></th>"
        if has_fact: hdr += "<th>Итого<br><small>Факт</small></th><th>Итого<br><small>Δ</small></th>"

        rows_html = ""
        for res in results:
            fc_vals = [round(res['fc_clean'].iloc[i],1) for i in range(steps)]
            sum_fc   = round(sum(fc_vals),1)
            sum_fact = round(sum(res['fact']),1)
            cells = (f"<td><b>{res['grp']}</b></td>"
                     f"<td><code>{res['article']}</code></td>"
                     f"<td style='text-align:left'>{res['name']}</td>"
                     f"<td>{badge_html(res['method_clean'])}</td>"
                     f"<td>{res['zero_pct']}%</td>")
            for i, v in enumerate(fc_vals):
                cells += f"<td><b>{v:,.1f}</b></td>"
                if has_fact:
                    cells += f"<td>{res['fact'][i]:,.1f}</td>"
                    cells += f"<td>{delta_html(round(res['fact'][i]-v,1))}</td>"
            cells += f"<td><b>{sum_fc:,.1f}</b></td>"
            if has_fact:
                cells += f"<td><b>{sum_fact:,.1f}</b></td>"
                cells += f"<td>{delta_html(round(sum_fact-sum_fc,1))}</td>"
            rows_html += f"<tr>{cells}</tr>"

        st.markdown(f'<div style="overflow-x:auto"><table class="fc-table">'
                    f'<thead><tr>{hdr}</tr></thead><tbody>{rows_html}</tbody></table></div>',
                    unsafe_allow_html=True)
        st.markdown("")

        # Графики
        st.markdown('<div class="sect-title">Графики</div>', unsafe_allow_html=True)
        ncols = min(2, len(results))
        cols_g = st.columns(ncols)
        for i, res in enumerate(results):
            with cols_g[i % ncols]:
                out_note = f" {len(res['outliers'])} выбр." if res['outliers'] else ""
                st.markdown(f"**Гр.{res['grp']}** — {res['name'][:40]}{out_note}")
                st.plotly_chart(make_chart(res, show_clean),
                                use_container_width=True,
                                config={'displayModeBar': False})
                if res['outliers']:
                    with st.expander(f"Выбросы ({len(res['outliers'])} шт.)"):
                        for o in res['outliers']:
                            st.markdown(f"**{o['date']}**: {o['original']} → "
                                        f"{o['replacement']} (порог {o['threshold']})")

render_tab(r_sale,   tab_sale)
render_tab(r_repair, tab_repair)

# ── СКАЧАТЬ ───────────────────────────────────────────────────────────────────
st.divider()
excel_bytes = build_excel_output(r_sale, r_repair, fc_months)
groups_str = '_'.join(str(g) for g in selected_groups[:5])
st.download_button(
    "Скачать Excel-отчёт",
    data=excel_bytes,
    file_name=f"forecast_{groups_str}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary", use_container_width=True,
)
