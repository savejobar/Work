import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_excel_output(
    results_sale: list[dict],
    results_repair: list[dict],
    fc_months: list[tuple[int, int]],
) -> bytes:
    """
    Строит Excel-отчёт с двумя листами: «Продажи» и «Ремонт».

    Args:
        results_sale:   Список словарей с результатами прогноза продаж.
                        Каждый словарь — один результат из run_forecast().
        results_repair: То же для ремонта.
        fc_months:      Список кортежей (год, месяц) прогнозного горизонта.

    Returns:
        Байты готового .xlsx файла. Передаётся напрямую в st.download_button.
    """
    from app.components import month_label

    wb   = openpyxl.Workbook()
    ws1  = wb.active
    ws1.title = "Продажи"
    ws2  = wb.create_sheet("Ремонт")

    _write_sheet(ws1, results_sale,   fc_months)
    _write_sheet(ws2, results_repair, fc_months)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    results: list[dict],
    fc_months: list[tuple[int, int]],
) -> None:
    """Заполняет один лист Excel данными прогноза."""
    from app.components import month_label

    thin = Side(style="thin", color="D0D7DE")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def w(
        r: int,
        c: int,
        v,
        bold: bool = False,
        bg: str | None = None,
        align: str = "center",
        fmt: str | None = None,
        color: str = "000000",
    ) -> None:
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="Calibri", bold=bold, color=color, size=10)
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=True
        )
        cell.border = brd
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if fmt:
            cell.number_format = fmt

    labels   = [month_label(y, m) for y, m in fc_months]
    has_fact = any(any(v > 0 for v in r["fact"]) for r in results)
    steps    = len(fc_months)

    # ── Заголовки ─────────────────────────────────────────────────────────
    headers = ["№", "Группа", "Артикул", "Номенклатура", "Метод", "Нули %"]
    for lbl in labels:
        headers += [f"{lbl} Прогноз", f"{lbl} Без выбр."]
        if has_fact:
            headers += [f"{lbl} Факт", f"{lbl} Δ"]
    headers += ["Итого Прогноз", "Итого Без выбр."]
    if has_fact:
        headers += ["Итого Факт", "Итого Δ"]

    for ci, h in enumerate(headers, 1):
        w(1, ci, h, bold=True, bg="1A2744", color="FFFFFF")
    ws.row_dimensions[1].height = 32

    col_widths = (
        [4, 8, 14, 40, 14, 8]
        + ([12, 14, 12, 10] if has_fact else [12, 14]) * len(labels)
        + [14, 16]
        + ([14, 10] if has_fact else [])
    )
    for ci, cw in enumerate(col_widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(ci)].width = cw

    # ── Строки данных ─────────────────────────────────────────────────────
    for ri, res in enumerate(results, 2):
        bg = "FFFFFF" if ri % 2 == 0 else "F0F4FA"
        ws.row_dimensions[ri].height = 20

        w(ri, 1, ri - 1, bg=bg)
        w(ri, 2, res["grp"],     bg=bg)
        w(ri, 3, res["article"], bg=bg)
        w(ri, 4, res["name"],    bg=bg, align="left")
        w(ri, 5, res["method_clean"], bg=bg)
        w(ri, 6, f"{res['zero_pct']}%", bg=bg)

        sum_raw = sum_clean = sum_fact = 0
        base = 7

        for i in range(steps):
            fc_r = round(res["fc_raw"].iloc[i],   1)
            fc_c = round(res["fc_clean"].iloc[i], 1)
            fact = res["fact"][i]
            sum_raw   += fc_r
            sum_clean += fc_c
            sum_fact  += fact

            w(ri, base,     fc_r, bg="FFF3CD", color="B8520A", bold=True, fmt="#,##0.0")
            w(ri, base + 1, fc_c, bg="F4ECF7", color="5B2C6F", bold=True, fmt="#,##0.0")

            if has_fact:
                w(ri, base + 2, fact, bg="D5F5E3", color="117A65", bold=True, fmt="#,##0.0")
                d  = round(fact - fc_c, 1)
                dv = f"{'+'  if d > 0 else ''}{d:.1f}"
                w(ri, base + 3, dv, bg=bg,
                  color=("15803D" if d >= 0 else "DC2626"), bold=True)
                base += 4
            else:
                base += 2

        w(ri, base,     round(sum_raw,   1), bg="FFF3CD", color="B8520A", bold=True, fmt="#,##0.0")
        w(ri, base + 1, round(sum_clean, 1), bg="F4ECF7", color="5B2C6F", bold=True, fmt="#,##0.0")

        if has_fact:
            w(ri, base + 2, round(sum_fact, 1), bg="D5F5E3", color="117A65", bold=True, fmt="#,##0.0")
            qd = round(sum_fact - sum_clean, 1)
            w(ri, base + 3,
              f"{'+'  if qd > 0 else ''}{qd:.1f}",
              bg=bg, color=("15803D" if qd >= 0 else "DC2626"), bold=True)

    ws.freeze_panes = "G2"
